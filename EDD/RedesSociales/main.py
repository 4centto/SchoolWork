import openpyxl as o #Modulo para abrir el archivo excel

#Metodo para obtener el mes pero en formato numerico
def calcularMes(mes):
    mes = str.lower(mes)
    if (mes == "enero"):
        return 1
    elif (mes == "febrero"):
        return 2
    elif (mes == "marzo"):
        return 3
    elif (mes == "abril"):
        return 4
    elif (mes == "mayo"):
        return 5
    elif (mes == "junio"):
        return 6
    else:
        return 0

file = openpyxl.load_workbook('redessociales.xlsx') #Abrir rchivo
hoja = file.active #Obtener la hoja activa

totalFilas = hoja.max_row #Obtener las filas y columnas
totalColumnas = hoja.max_column

#Impresion de la tabla completa
print("")
print(" ** TABLA ** ")
print("")
for i in range(1, totalFilas + 1):
    for j in range(1, totalColumnas + 1):
        celda = hoja.cell(row = i, column = j)
        print("[", celda.value, "]", end="")

    print("")

print("\n\n")

#Obtencion de los seguidores para calcular finalmente su diferencia
seguidoresEnero = hoja.cell(row=9, column=4).value
seguidoresJunio = hoja.cell(row=9, column=9).value
difEneroJunio = seguidoresJunio - seguidoresEnero
print("* La diferencia de seguidores en twitter de Enero a Junio es de:", difEneroJunio)
print("")

#Calcular la diferencia de visualizaciones de youtube
print(" * DIFERENCIA DE VIZUALIZACIONES EN YOUTUBE * ")
mesUno = input("Dame el primer mes: ")
mesDos = input("Dame el segundo mes: ")

print(" -> La diferencia de", mesUno, "a", mesDos," es de: \n")

mesUno = calcularMes(mesUno) + 3 #Obtener el mes en formato numerico
mesDos = calcularMes(mesDos) + 3

valUno = hoja.cell(row=17, column=mesUno).value
valDos = hoja.cell(row=17, column=mesDos).value
resultado = valDos - valUno

if resultado < 0:
    resultado = resultado * -1

print("  ", resultado, "visulaizaciones <-")

#Promedios de crecimiento de Twitter y facebook respectivamente
print("\n\n")
print(" ** PROMEDIO DE CRECIMIENTO DE TWITTER DE ENERO A JUNIO ** ")
acum = 0

for i in range(1, totalColumnas + 1):
    if i >= 4 and i <= 9:
        acum += hoja.cell(row=10, column=i).value

print(" * El promedio es: ", round(acum / 6))

print("\n ** PROMEDIO DE CRECIMIENTO DE FACEBOOK DE ENERO A JUNIO ** ")
acum = 0
for i in range(1, totalColumnas + 1):
    if i >= 4 and i <= 9:
        acum += hoja.cell(row=3, column=i).value

print(" * El promedio es: ", round(acum / 6))